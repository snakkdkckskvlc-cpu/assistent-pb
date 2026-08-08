"""Чтение смет и актов из XLSX.

Проверяется не «файл прочитался», а то, что программа правильно ПОНЯЛА файл:
где шапка, какая колонка что значит и какие строки позициями не являются.
Ошибка здесь не видна в отчёте — он просто окажется посчитан не по тем
колонкам и будет выглядеть правильным.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from fire_safety_backend.infrastructure.parsers.xlsx_parser import read_table, sheet_names
from openpyxl import Workbook


def _book(rows: list[list[object]], path: Path, *, title: str = "Смета") -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = title
    for row in rows:
        ws.append(row)
    wb.save(path)
    return path


# --- типовая смета ---

СМЕТА = [
    ["ЛОКАЛЬНАЯ СМЕТА № 15"],
    ["Объект: цех №2 ПАО НЛМК"],
    [],
    ["№", "Наименование работ и материалов", "Ед. изм.", "Количество", "Цена", "Сумма"],
    [1, "Кабель ВВГнг 3х1,5", "м", 100, 50, 5000],
    [2, "Извещатель ИП 212-45", "шт", 20, 600, 12000],
    [3, "Монтаж трубопровода", "м", 45.5, 200, 9100],
    ["", "ИТОГО", "", "", "", 26100],
]


def test_header_and_columns_are_recognised(tmp_path: Path) -> None:
    разбор = read_table(_book(СМЕТА, tmp_path / "смета.xlsx"))
    assert разбор.шапка_найдена
    assert разбор.строка_шапки == 3  # с нуля: четвёртая строка файла
    assert разбор.колонки["название"] == 2
    assert разбор.колонки["количество"] == 4
    assert разбор.колонки["сумма"] == 6
    assert разбор.колонки["цена"] == 5


def test_positions_are_read(tmp_path: Path) -> None:
    разбор = read_table(_book(СМЕТА, tmp_path / "смета.xlsx"))
    assert [r.название for r in разбор.строки] == [
        "Кабель ВВГнг 3х1,5",
        "Извещатель ИП 212-45",
        "Монтаж трубопровода",
    ]
    assert разбор.строки[0].количество == 100_000  # тысячные доли
    assert разбор.строки[0].сумма == 500_000  # копейки
    assert разбор.строки[2].количество == 45_500  # 45,5 — дробное количество


def test_total_row_is_not_a_position(tmp_path: Path) -> None:
    """Строку «ИТОГО» нельзя считать позицией: она равна сумме строк выше,
    и в сравнении удвоила бы итог файла."""
    разбор = read_table(_book(СМЕТА, tmp_path / "смета.xlsx"))
    assert разбор.пропущено_итогов == 1
    assert all("итого" not in r.название.casefold() for r in разбор.строки)
    assert sum(r.сумма or 0 for r in разбор.строки) == 2_610_000


@pytest.mark.parametrize(
    "мусор",
    ["Итого:", "ВСЕГО по смете", "в том числе НДС", "НДС 20%", "К оплате", "Раздел 1. Монтаж"],
)
def test_service_rows_are_skipped(tmp_path: Path, мусор: str) -> None:
    rows = [["Наименование", "Количество", "Сумма"], ["Кабель", 1, 100], [мусор, "", 100]]
    разбор = read_table(_book(rows, tmp_path / "x.xlsx"))
    assert [r.название for r in разбор.строки] == ["Кабель"]


# --- разные написания шапки ---


@pytest.mark.parametrize(
    "шапка",
    [
        ["Наименование", "Кол-во", "Сумма"],
        ["Название позиции", "Количество", "Стоимость"],
        ["Номенклатура", "Объем", "Сумма, руб."],
        ["Наименование работ", "Кол.", "Всего"],
        ["Товар", "Количество, ед.", "Сумма с НДС"],
    ],
)
def test_header_synonyms(tmp_path: Path, шапка: list[str]) -> None:
    разбор = read_table(_book([шапка, ["Кабель", 5, 500]], tmp_path / "x.xlsx"))
    assert разбор.шапка_найдена, f"шапка не опознана: {шапка}"
    assert разбор.колонки.get("количество") == 2, f"количество не найдено в {шапка}"
    assert разбор.колонки.get("сумма") == 3, f"сумма не найдена в {шапка}"
    assert разбор.строки[0].количество == 5000


def test_sum_wins_over_price_for_stoimost(tmp_path: Path) -> None:
    """«Стоимость» — это сумма, а не цена за единицу: сначала ищется сумма."""
    разбор = read_table(
        _book([["Наименование", "Цена", "Стоимость"], ["Кабель", 50, 5000]], tmp_path / "x.xlsx")
    )
    assert разбор.колонки["сумма"] == 3
    assert разбор.колонки["цена"] == 2
    assert разбор.строки[0].сумма == 500_000


# --- когда шапки нет ---


def test_no_header_falls_back_and_warns(tmp_path: Path) -> None:
    """Без шапки колонки — догадка, и человек обязан об этом узнать.

    Молча посчитать по угаданной колонке и выдать отчёт — худший исход:
    он выглядит правильным.
    """
    rows = [["Кабель ВВГнг 3х1,5", 100, 5000], ["Извещатель ИП 212-45", 20, 12000]]
    разбор = read_table(_book(rows, tmp_path / "без_шапки.xlsx"))
    assert not разбор.шапка_найдена
    assert разбор.предупреждения, "человека обязаны предупредить"
    assert "шапка" in разбор.предупреждения[0].casefold()
    assert [r.название for r in разбор.строки] == ["Кабель ВВГнг 3х1,5", "Извещатель ИП 212-45"]
    assert разбор.строки[0].сумма == 500_000


def test_missing_quantity_column_warns(tmp_path: Path) -> None:
    разбор = read_table(_book([["Наименование", "Сумма"], ["Кабель", 5000]], tmp_path / "x.xlsx"))
    assert разбор.шапка_найдена
    assert any("количеств" in w.casefold() for w in разбор.предупреждения)
    assert разбор.строки[0].количество is None


def test_empty_sheet_warns_and_does_not_crash(tmp_path: Path) -> None:
    разбор = read_table(_book([["Наименование", "Количество", "Сумма"]], tmp_path / "x.xlsx"))
    assert разбор.строки == []
    assert any("ни одной строки" in w for w in разбор.предупреждения)


# --- формулы, листы, мелочи реальных файлов ---


def test_formula_values_are_read_not_formulas(tmp_path: Path) -> None:
    """В смете сумма почти всегда формула. Без data_only в ячейке приедет
    строка «=D5*E5», которую не с чем сравнивать."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Наименование", "Количество", "Сумма"])
    ws.append(["Кабель", 100, "=B2*50"])
    path = tmp_path / "формулы.xlsx"
    wb.save(path)
    разбор = read_table(path)
    # openpyxl без пересчёта отдаёт None вместо значения незакэшированной
    # формулы — это НЕ повод молча подставить ноль.
    assert разбор.строки[0].сумма is None


def test_second_sheet_can_be_chosen(tmp_path: Path) -> None:
    wb = Workbook()
    wb.active.title = "Титул"
    wb.active.append(["Просто текст"])
    ws2 = wb.create_sheet("Смета")
    ws2.append(["Наименование", "Количество", "Сумма"])
    ws2.append(["Кабель", 1, 100])
    path = tmp_path / "две.xlsx"
    wb.save(path)

    assert sheet_names(path) == ["Титул", "Смета"]
    разбор = read_table(path, sheet="Смета")
    assert разбор.лист == "Смета"
    assert [r.название for r in разбор.строки] == ["Кабель"]


def test_blank_rows_inside_table_are_skipped(tmp_path: Path) -> None:
    rows = [
        ["Наименование", "Количество", "Сумма"],
        ["Кабель", 1, 100],
        [None, None, None],
        ["Труба 25", 2, 200],
    ]
    разбор = read_table(_book(rows, tmp_path / "x.xlsx"))
    assert [r.название for r in разбор.строки] == ["Кабель", "Труба 25"]


def test_row_numbers_point_at_the_file(tmp_path: Path) -> None:
    """Номер строки нужен, чтобы человек нашёл расхождение в своём файле."""
    разбор = read_table(_book(СМЕТА, tmp_path / "смета.xlsx"))
    assert разбор.строки[0].номер == 5  # пятая строка файла, считая с единицы
